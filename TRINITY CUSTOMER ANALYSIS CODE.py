"""
Customer Report Generator (FINAL TEMPLATE)

Creates ONE Excel workbook with these sheets:
1) Total_Pivot_Table  (Item x YYYY-MM, totals, colored by status)
2) Alert_30_vs_3M     (Last 30 vs prev 3 months)   baseline = AVG monthly qty over prior 3 months
3) Alert_30_vs_6M     (Last 30 vs prev 6 months)   baseline = AVG monthly qty over prior 6 months
4) Alert_30_vs_1Y     (Last 30 vs prev 12 months)  baseline = AVG monthly qty over prior 12 months
5) Alert_30_vs_All    (Last 30 vs all history)     baseline = AVG monthly qty over all prior months
6) Master_Alert_Sheet (ALL items, baseline avgs + alerts across 3M/6M/1Y/All; includes Days_Since_Last_Purchase)
7) One_Time_Buy_Items (items purchased only once in last 1 year)
8) Regular_Buy_Items  (frequent items + Avg Monthly Quantity)

Also:
- Converts each sheet range to an Excel Table (Ctrl+T style)
- Conditional formatting on Alert column (STOPPED red, DECREASE amber, INCREASE/NEW green)
- Pivot rows colored based on worst alert across Master sheet

Usage:
  pip install pandas openpyxl
  python shopping_dashboard_template_final.py --input "file.csv" --output "report.xlsx"
"""

import argparse
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import FormulaRule


# --------- Colors ----------
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")     # STOPPED
AMBER = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")   # DECREASE
GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # INCREASE / NEW ITEM


def find_col(cols, candidates):
    norm = {c.strip().lower(): c for c in cols}
    for cand in candidates:
        k = cand.strip().lower()
        if k in norm:
            return norm[k]
    raise ValueError(f"Missing required column. Tried {candidates}. Found {list(cols)}")


def monthly_avg_between(w: pd.DataFrame, start: pd.Timestamp, end: pd.Timestamp) -> pd.Series:
    """Average MONTHLY quantity per item within [start, end]."""
    d = w[(w["Date"] >= start) & (w["Date"] <= end)].copy()
    if d.empty:
        return pd.Series(dtype=float)
    d["YYYY-MM"] = d["Date"].dt.to_period("M").astype(str)
    m = d.groupby(["Item Description", "YYYY-MM"])["Quantity"].sum().reset_index()
    return m.groupby("Item Description")["Quantity"].mean()


def last_purchase_month_between(w: pd.DataFrame, start: pd.Timestamp, end: pd.Timestamp, item: str) -> str:
    d = w[(w["Item Description"] == item) & (w["Date"] >= start) & (w["Date"] <= end)]
    if d.empty:
        return ""
    return d["Date"].dt.to_period("M").astype(str).max()


def classify(base_avg: float, last30_qty: float):
    """Compare Last 30-day qty to baseline avg monthly qty."""
    if base_avg == 0 and last30_qty > 0:
        return "NEW ITEM", "New purchase in last 30 days"
    if base_avg > 0 and last30_qty == 0:
        return "STOPPED", "No purchase in last 30 days (stopped after baseline)"
    if base_avg > 0 and last30_qty <= base_avg * 0.5:
        return "DECREASE", "Reduced vs baseline average"
    if base_avg > 0 and last30_qty >= base_avg * 1.5:
        return "INCREASE", "Increased vs baseline average"
    return "", ""


def build_alert(work: pd.DataFrame, as_of: pd.Timestamp, months):
    """months: int (3/6/12) or 'ALL'"""
    last30_start = as_of - pd.Timedelta(days=29)

    last30 = work[(work["Date"] >= last30_start) & (work["Date"] <= as_of)]
    l30_qty = last30.groupby("Item Description")["Quantity"].sum()

    if months == "ALL":
        base_start = work["Date"].min().normalize()
        base_type = "ALL"
    else:
        base_start = (last30_start - pd.DateOffset(months=int(months))).normalize()
        base_type = f"{int(months)}M"
    base_end = (last30_start - pd.Timedelta(days=1)).normalize()

    base_avg = monthly_avg_between(work, base_start, base_end)
    last_date = work.groupby("Item Description")["Date"].max()

    rows = []
    items = set(base_avg.index).union(l30_qty.index)
    for item in sorted(items):
        b = float(base_avg.get(item, 0))
        q30 = float(l30_qty.get(item, 0))
        alert, reason = classify(b, q30)
        if not alert:
            continue

        ld = last_date.get(item)
        days_since = (as_of - ld.normalize()).days if pd.notna(ld) else ""

        if alert == "STOPPED":
            exact_month = last_purchase_month_between(work, base_start, base_end, item)
        else:
            exact_month = last_purchase_month_between(work, last30_start, as_of, item) or last_purchase_month_between(work, base_start, base_end, item)

        rows.append([
            item,
            round(b, 2),
            q30,
            alert,
            reason,
            exact_month,
            days_since,
            base_type,
            as_of.date()
        ])

    return pd.DataFrame(rows, columns=[
        "Item Description",
        "Baseline_Avg_Monthly_Qty",
        "Qty_Last30",
        "Alert",
        "Change_Reason",
        "Exact_Stop_or_Reduce_Month",
        "Days_Since_Last_Purchase",
        "Baseline_Type",
        "As_Of"
    ])


def build_master(work: pd.DataFrame, as_of: pd.Timestamp):
    last30_start = as_of - pd.Timedelta(days=29)
    items_all = sorted(work["Item Description"].unique())

    last30_qty = work[(work["Date"] >= last30_start) & (work["Date"] <= as_of)].groupby("Item Description")["Quantity"].sum()
    last_date = work.groupby("Item Description")["Date"].max()

    def base_avg_for(months):
        if months == "ALL":
            bs = work["Date"].min().normalize()
        else:
            bs = (last30_start - pd.DateOffset(months=int(months))).normalize()
        be = (last30_start - pd.Timedelta(days=1)).normalize()
        return monthly_avg_between(work, bs, be)

    avg3 = base_avg_for(3)
    avg6 = base_avg_for(6)
    avg12 = base_avg_for(12)
    avgA = base_avg_for("ALL")

    rows = []
    for item in items_all:
        q30 = float(last30_qty.get(item, 0))
        ld = last_date.get(item)
        days_since = (as_of - ld.normalize()).days if pd.notna(ld) else ""

        b3 = float(avg3.get(item, 0))
        b6 = float(avg6.get(item, 0))
        b12 = float(avg12.get(item, 0))
        bA = float(avgA.get(item, 0))

        a3, _ = classify(b3, q30)
        a6, _ = classify(b6, q30)
        a12, _ = classify(b12, q30)
        aA, _ = classify(bA, q30)

        has_any = "YES" if any([a3, a6, a12, aA]) else "NO"

        rows.append([
            item, as_of.date(), last30_start.date(), as_of.date(),
            q30,
            round(b3, 2), a3,
            round(b6, 2), a6,
            round(b12, 2), a12,
            round(bA, 2), aA,
            days_since,
            has_any
        ])

    master = pd.DataFrame(rows, columns=[
        "Item Description","As_Of","Last30_Start","Last30_End",
        "Qty_Last30",
        "Baseline_Avg_3M","Alert_3M",
        "Baseline_Avg_6M","Alert_6M",
        "Baseline_Avg_1Y","Alert_1Y",
        "Baseline_Avg_All","Alert_ALL",
        "Days_Since_Last_Purchase",
        "Has_Any_Alert"
    ]).sort_values(["Has_Any_Alert","Item Description"], ascending=[False, True]).reset_index(drop=True)

    return master


def build_one_time(work: pd.DataFrame, as_of: pd.Timestamp):
    start_1y = as_of - pd.Timedelta(days=364)
    ly = work[(work["Date"] >= start_1y) & (work["Date"] <= as_of)].copy()
    one_time = (
        ly.groupby("Item Description", as_index=False)
          .agg(
              Unique_Purchase_Days=("Date", lambda x: x.dt.date.nunique()),
              Total_Quantity=("Quantity","sum"),
              Last_Purchase_Date=("Date","max"),
          )
    )
    one_time = one_time[one_time["Unique_Purchase_Days"] == 1].copy()
    one_time["Last_Purchase_Month"] = one_time["Last_Purchase_Date"].dt.to_period("M").astype(str)
    one_time.drop(columns=["Last_Purchase_Date"], inplace=True)
    return one_time.sort_values(["Total_Quantity","Item Description"], ascending=[False, True]).reset_index(drop=True)


def build_regular(work: pd.DataFrame):
    w = work.copy()
    w["YYYY-MM"] = w["Date"].dt.to_period("M").astype(str)
    monthly = (
        w.groupby(["Item Description","YYYY-MM"], as_index=False)
         .agg(Monthly_Quantity=("Quantity","sum"))
    )
    regular = (
        monthly.groupby("Item Description", as_index=False)
               .agg(
                   Months_Purchased=("YYYY-MM","nunique"),
                   Avg_Monthly_Quantity=("Monthly_Quantity","mean"),
                   Total_Quantity=("Monthly_Quantity","sum"),
               )
    )
    regular = regular[(regular["Months_Purchased"] >= 3) & (regular["Avg_Monthly_Quantity"] > 0)].copy()
    regular["Avg_Monthly_Quantity"] = regular["Avg_Monthly_Quantity"].round(2)
    return regular.sort_values(["Months_Purchased","Avg_Monthly_Quantity"], ascending=[False, False]).reset_index(drop=True)


def build_pivot(work: pd.DataFrame):
    w = work.copy()
    w["YYYY-MM"] = w["Date"].dt.to_period("M").astype(str)
    pivot = w.pivot_table(index="Item Description", columns="YYYY-MM", values="Quantity", aggfunc="sum", fill_value=0)
    pivot = pivot.reindex(sorted(pivot.columns), axis=1)
    pivot["TOTAL"] = pivot.sum(axis=1)
    total_row = pivot.sum(axis=0)
    total_row.name = "TOTAL"
    return pd.concat([pivot, total_row.to_frame().T]).reset_index()


def add_table(ws):
    if ws.max_row < 2 or ws.max_column < 2:
        return
    end_col_letter = ws.cell(row=1, column=ws.max_column).column_letter
    ref = f"A1:{end_col_letter}{ws.max_row}"
    name = ws.title.replace(" ", "_").replace("-", "_")[:30]
    ws._tables.clear()
    t = Table(displayName=name, ref=ref)
    t.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(t)


def auto_fit(ws, max_width=45):
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, max_width)


def add_alert_cf(ws):
    headers = [c.value for c in ws[1]]
    if "Alert" not in headers or ws.max_row < 2:
        return
    idx = headers.index("Alert") + 1
    col_letter = ws.cell(row=1, column=idx).column_letter
    rng = f"{col_letter}2:{col_letter}{ws.max_row}"
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'${col_letter}2="STOPPED"'], fill=RED))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'${col_letter}2="DECREASE"'], fill=AMBER))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'OR(${col_letter}2="INCREASE",${col_letter}2="NEW ITEM")'], fill=GREEN))


def color_pivot_rows(pivot_ws, master_df: pd.DataFrame):
    priority = {"STOPPED": 3, "DECREASE": 2, "INCREASE": 1, "NEW ITEM": 1, "": 0}
    status_map = {}
    for _, r in master_df.iterrows():
        statuses = [r["Alert_3M"], r["Alert_6M"], r["Alert_1Y"], r["Alert_ALL"]]
        best = ""
        best_p = -1
        for s in statuses:
            p = priority.get(s, 0)
            if p > best_p:
                best_p = p
                best = s
        status_map[r["Item Description"]] = best

    for row in range(2, pivot_ws.max_row + 1):
        item = pivot_ws.cell(row=row, column=1).value
        if item is None or item == "TOTAL":
            continue
        s = status_map.get(str(item), "")
        fill = None
        if s == "STOPPED":
            fill = RED
        elif s == "DECREASE":
            fill = AMBER
        elif s in ("INCREASE", "NEW ITEM"):
            fill = GREEN
        if fill:
            for c in range(1, pivot_ws.max_column + 1):
                pivot_ws.cell(row=row, column=c).fill = fill


def build_output(input_csv: Path, output_xlsx: Path):
    df = pd.read_csv(input_csv)

    desc_col = find_col(df.columns, ["description", "item description", "item", "product", "name"])
    qty_col  = find_col(df.columns, ["qty", "quantity"])
    date_col = find_col(df.columns, ["date"])

    work = df[[desc_col, qty_col, date_col]].copy()
    work.columns = ["Item Description", "Quantity", "Date"]
    work["Date"] = pd.to_datetime(work["Date"], errors="coerce")
    work = work.dropna(subset=["Date"]).copy()
    work["Quantity"] = pd.to_numeric(work["Quantity"], errors="coerce").fillna(0)
    work["Item Description"] = work["Item Description"].astype(str)

    as_of = work["Date"].max().normalize()

    pivot_df = build_pivot(work)
    alert_3m = build_alert(work, as_of, 3)
    alert_6m = build_alert(work, as_of, 6)
    alert_1y = build_alert(work, as_of, 12)
    alert_all = build_alert(work, as_of, "ALL")
    master_df = build_master(work, as_of)
    one_time_df = build_one_time(work, as_of)
    regular_df = build_regular(work)

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        pivot_df.to_excel(writer, index=False, sheet_name="Total_Pivot_Table")
        alert_3m.to_excel(writer, index=False, sheet_name="Alert_30_vs_3M")
        alert_6m.to_excel(writer, index=False, sheet_name="Alert_30_vs_6M")
        alert_1y.to_excel(writer, index=False, sheet_name="Alert_30_vs_1Y")
        alert_all.to_excel(writer, index=False, sheet_name="Alert_30_vs_All")
        master_df.to_excel(writer, index=False, sheet_name="Master_Alert_Sheet")
        one_time_df.to_excel(writer, index=False, sheet_name="One_Time_Buy_Items")
        regular_df.to_excel(writer, index=False, sheet_name="Regular_Buy_Items")

    wb = load_workbook(output_xlsx)

    # basic format + table + CF
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
        auto_fit(ws)
        add_table(ws)
        if ws.title.startswith("Alert_30_vs_") or ws.title == "Master_Alert_Sheet":
            add_alert_cf(ws)

    color_pivot_rows(wb["Total_Pivot_Table"], master_df)
    wb.save(output_xlsx)


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--input", required=True, help="Input CSV path")
    p.add_argument("--output", required=True, help="Output XLSX path")
    args = p.parse_args()
    build_output(Path(args.input), Path(args.output))
    print("DONE ✅")


if __name__ == "__main__":
    main()
